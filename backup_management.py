"""
SPDX-License-Identifier: GPL-3.0-or-later

Copyright (C) 2026 Al Gelders

This file is part of the airscenting an trailing logging programs

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU General Public License for more details.

You should have received a copy of the GNU General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
"""

"""
Backup Management Module for Air-Scenting Logger

Provides shared utilities for backup operations:
- UUID generation for session tracking
- Update time management
- Sync operations between database and JSON folders

This module is designed to be shared between Airscent and Trailing tabs.

Author: AI Assistant
Date: 2025-01-07
"""

import uuid
import json
import shutil
import threading
from datetime import datetime
from pathlib import Path


def generate_session_uuid():
    """
    Generate a unique UUID for a new session.
    
    This UUID is used to track sessions across the database and JSON backup files,
    enabling synchronization between primary and secondary backup locations.
    
    Returns:
        str: A new UUID4 string (e.g., 'a1b2c3d4-e5f6-7890-abcd-ef1234567890')
    
    Note:
        - UUID should only be generated when creating a NEW session (Save Session)
        - UUID should NOT be regenerated when updating an existing session (Update Session)
    """
    return str(uuid.uuid4())


def get_current_update_time():
    """
    Get the current datetime for update_time field.
    
    Returns:
        datetime: Current UTC datetime
        
    Note:
        This should be called on EVERY save or update operation.
    """
    return datetime.utcnow()


def get_update_time_iso():
    """
    Get the current datetime as ISO format string for JSON serialization.
    
    Returns:
        str: Current datetime in ISO format (e.g., '2025-01-07T14:30:00.123456')
    """
    return datetime.utcnow().isoformat()


# ============================================================================
# SYNC UTILITIES
# ============================================================================

def _format_update_time(update_time):
    """
    Safely format update_time to ISO string.
    Handles both datetime objects and strings.
    """
    if not update_time:
        return ""
    if isinstance(update_time, str):
        return update_time
    try:
        return update_time.isoformat()
    except:
        return str(update_time)


def _parse_update_time(update_time_str):
    """
    Safely parse update_time string to datetime.
    Returns None if parsing fails.
    """
    if not update_time_str:
        return None
    if isinstance(update_time_str, datetime):
        return update_time_str
    try:
        return datetime.fromisoformat(update_time_str)
    except:
        return None


class BackupSyncManager:
    """
    Manages synchronization between database and JSON backup folders.
    
    Handles:
    - Startup sync between DB, primary JSON, and secondary JSON
    - Background thread execution
    - Blocking Edit/Hide operations during sync
    """
    
    def __init__(self):
        self.sync_in_progress = False
        self.sync_thread = None
        self.sync_results = {
            "db_to_json": 0,
            "json_to_db": 0,
            "primary_to_secondary": 0,
            "secondary_to_primary": 0,
            "errors": []
        }
    
    def is_sync_in_progress(self):
        """Check if sync is currently running"""
        return self.sync_in_progress
    
    def start_background_sync(self, db_type, primary_folder, secondary_folder, 
                               on_complete=None, status_callback=None):
        """
        Start background sync in a separate thread.
        
        Args:
            db_type: Database type (sqlite, postgres, etc.)
            primary_folder: Path to primary JSON folder
            secondary_folder: Path to secondary JSON folder (can be None)
            on_complete: Callback function when sync completes (receives sync_results)
            status_callback: Callback to update status bar (receives message string)
        """
        if self.sync_in_progress:
            # print("Sync already in progress")
            return False
        
        self.sync_in_progress = True
        self.sync_results = {
            "db_to_json": 0,
            "json_to_db": 0,
            "primary_to_secondary": 0,
            "secondary_to_primary": 0,
            "errors": []
        }
        
        def sync_worker():
            try:
                if status_callback:
                    status_callback("Sync: Scanning JSON folders...")
                
                import time
                self._run_sync(db_type, primary_folder, secondary_folder, status_callback)
                
            except Exception as e:
                self.sync_results["errors"].append(f"Sync error: {e}")
                # print(f"Sync error: {e}")
                import traceback
                traceback.print_exc()
            finally:
                self.sync_in_progress = False
                if on_complete:
                    on_complete(self.sync_results)
        
        self.sync_thread = threading.Thread(target=sync_worker, daemon=True)
        self.sync_thread.start()
        return True
    
    def _run_sync(self, db_type, primary_folder, secondary_folder, status_callback):
        """
        Main sync logic - runs in background thread.
        
        Sync order:
        1. Scan primary JSON folder
        2. Scan secondary JSON folder (if exists)
        3. Get DB sessions with UUID/update_time
        4. Sync DB ÃƒÂ¢Ã¢â‚¬Â Ã¢â‚¬â„¢ Primary JSON (DB newer or missing in JSON)
        5. Sync Primary JSON ÃƒÂ¢Ã¢â‚¬Â Ã¢â‚¬â„¢ DB (JSON newer)
        6. Sync Primary ÃƒÂ¢Ã¢â‚¬Â Ã¢â‚¬â„¢ Secondary (Primary newer or missing)
        7. Sync Secondary ÃƒÂ¢Ã¢â‚¬Â Ã¢â‚¬â„¢ Primary (Secondary newer, also update DB)
        """
        primary_path = Path(primary_folder) if primary_folder else None
        secondary_path = Path(secondary_folder) if secondary_folder else None
        
        # Step 1: Scan primary JSON folder
        primary_dict = {}
        if primary_path and primary_path.exists():
            if status_callback:
                status_callback("Sync: Scanning primary JSON folder...")
            primary_dict = scan_json_folder(primary_path)
            # print(f"Sync: Found {len(primary_dict)} sessions in primary JSON")
            pass
        
        # Step 2: Scan secondary JSON folder
        secondary_dict = {}
        if secondary_path and secondary_path.exists():
            if status_callback:
                status_callback("Sync: Scanning secondary JSON folder...")
            secondary_dict = scan_json_folder(secondary_path)
            # print(f"Sync: Found {len(secondary_dict)} sessions in secondary JSON")
            pass
        
        # Step 3: Get DB sessions
        if status_callback:
            status_callback("Sync: Reading database sessions...")
        db_sessions = get_db_sessions_for_sync(db_type)
        # print(f"Sync: Found {len(db_sessions)} sessions in database")
        pass
        
        # Step 4: Sync DB ÃƒÂ¢Ã¢â‚¬Â Ã¢â‚¬â„¢ Primary JSON
        if primary_path and primary_path.exists():
            if status_callback:
                status_callback("Sync: Updating JSON from database...")
            count = sync_db_to_json(db_sessions, primary_dict, primary_path, db_type)
            self.sync_results["db_to_json"] = count
            if count > 0:
                # Re-scan primary after updates
                primary_dict = scan_json_folder(primary_path)
        
        # Step 5: Sync Primary JSON ÃƒÂ¢Ã¢â‚¬Â Ã¢â‚¬â„¢ DB
        if primary_dict:
            if status_callback:
                status_callback("Sync: Updating database from JSON...")
            count = sync_json_to_db(primary_dict, db_sessions, db_type)
            self.sync_results["json_to_db"] = count
        
        # Step 6: Sync Primary ÃƒÂ¢Ã¢â‚¬Â Ã¢â‚¬â„¢ Secondary
        if secondary_path and secondary_path.exists() and primary_dict:
            if status_callback:
                status_callback("Sync: Mirroring to secondary backup...")
            count = sync_primary_to_secondary(primary_dict, secondary_dict, secondary_path)
            self.sync_results["primary_to_secondary"] = count
        
        # Step 7: Sync Secondary ÃƒÂ¢Ã¢â‚¬Â Ã¢â‚¬â„¢ Primary (and DB)
        if secondary_path and secondary_path.exists() and primary_path and primary_path.exists():
            if status_callback:
                status_callback("Sync: Checking secondary for newer files...")
            # Re-scan primary to get current state
            primary_dict = scan_json_folder(primary_path)
            count = sync_secondary_to_primary(secondary_dict, primary_dict, primary_path, db_type)
            self.sync_results["secondary_to_primary"] = count
        
        # print(f"Sync complete: {self.sync_results}")
        pass


def scan_json_folder(folder_path):
    """
    Scan a JSON folder and build a dictionary of sessions by UUID.
    
    Args:
        folder_path: Path to JSON folder
        
    Returns:
        dict: {uuid: {"update_time": datetime, "file_mtime": datetime, "filepath": Path, "data": dict}}
              Sessions without UUID are indexed by filename
    """
    result = {}
    folder = Path(folder_path)
    
    if not folder.exists():
        return result
    
    for json_file in folder.glob("*session*.json"):
        try:
            with open(json_file, 'r') as f:
                data = json.load(f)
            
            session_uuid = data.get("uuid")
            update_time_str = data.get("update_time")
            
            # Parse update_time from JSON
            update_time = None
            if update_time_str:
                try:
                    update_time = datetime.fromisoformat(update_time_str)
                except:
                    pass
            
            # Get file modification time as fallback
            file_mtime = datetime.fromtimestamp(json_file.stat().st_mtime)
            
            # If no update_time in JSON, use file modification time
            if update_time is None:
                update_time = file_mtime
            
            # Use UUID as key if available, otherwise use filename
            key = session_uuid if session_uuid else json_file.stem
            
            result[key] = {
                "update_time": update_time,
                "file_mtime": file_mtime,
                "filepath": json_file,
                "data": data,
                "has_uuid": bool(session_uuid)
            }
            
        except Exception as e:
            # print(f"Warning: Could not read {json_file}: {e}")
            pass
    
    return result


def get_db_sessions_for_sync(db_type):
    """
    Get all sessions from database with UUID and update_time.
    Includes both airscenting (training_sessions) and trailing (t_training_sessions) tables.
    
    Returns:
        dict: {uuid: {"update_time": datetime, "session_number": int, "dog_name": str, "data": dict}}
    """
    result = {}
    
    try:
        import config
        from sqlalchemy import text
        
        # Temporarily switch to correct DB type
        old_db_type = config.DB_TYPE
        config.DB_TYPE = db_type
        
        from database import engine, get_connection
        from importlib import reload
        import database
        
        if old_db_type != db_type:
            engine.dispose()
            reload(database)
        
        with database.get_connection() as conn:
            # Get airscenting sessions
            query = text("""
                SELECT id, session_number, dog_name, date, handler, session_purpose, 
                       field_support, location, search_area_size, num_subjects, 
                       handler_knowledge, weather, temperature, wind_direction, 
                       wind_speed, search_type, drive_level, subjects_found, 
                       comments, image_files, entry_type, update_time, uuid
                FROM training_sessions
                WHERE uuid IS NOT NULL AND uuid != ''
            """)
            rows = conn.execute(query).fetchall()
            
            for row in rows:
                session_uuid = row[22]  # uuid column
                update_time_raw = row[21]   # update_time column
                
                # Parse update_time to datetime object for consistent comparison
                update_time = _parse_update_time(update_time_raw)
                
                if session_uuid:
                    result[session_uuid] = {
                        "update_time": update_time,
                        "session_number": row[1],
                        "dog_name": row[2],
                        "data": {
                            "id": row[0],
                            "session_number": row[1],
                            "dog_name": row[2],
                            "date": str(row[3]) if row[3] else "",
                            "handler": row[4] or "",
                            "session_purpose": row[5] or "",
                            "field_support": row[6] or "",
                            "location": row[7] or "",
                            "search_area_size": row[8] or "",
                            "num_subjects": row[9] or "",
                            "handler_knowledge": row[10] or "",
                            "weather": row[11] or "",
                            "temperature": row[12] or "",
                            "wind_direction": row[13] or "",
                            "wind_speed": row[14] or "",
                            "search_type": row[15] or "",
                            "drive_level": row[16] or "",
                            "subjects_found": row[17] or "",
                            "comments": row[18] or "",
                            "image_files": row[19] or "",
                            "entry_type": row[20] or "",
                            "update_time": _format_update_time(update_time),
                            "uuid": session_uuid
                        }
                    }
            
            # Get trailing sessions
            try:
                t_query = text("""
                    SELECT id, t_session_number, t_dog_name, t_date, t_handler, t_field_support,
                           t_location, t_start_time, t_finish_time, t_trail_age, t_trail_length,
                           t_difficulty, t_trail_layer, t_cross_track_layer, t_cross_track_age,
                           t_weather_laying, t_temperature_laying, t_wind_speed_laying, t_wind_direction_laying, t_humidity_laying,
                           t_weather_running, t_temperature_running, t_wind_speed_running, t_wind_direction_running, t_humidity_running,
                           t_start_behavior, t_consistency, t_head_position, t_pace, t_indication,
                           t_time_to_complete, t_success_rate, t_impression, t_map_files,
                           update_time, uuid
                    FROM t_training_sessions
                    WHERE uuid IS NOT NULL AND uuid != ''
                """)
                t_rows = conn.execute(t_query).fetchall()
                
                for row in t_rows:
                    session_uuid = row[35]  # uuid column
                    update_time_raw = row[34]   # update_time column
                    
                    update_time = _parse_update_time(update_time_raw)
                    
                    if session_uuid:
                        result[session_uuid] = {
                            "update_time": update_time,
                            "session_number": row[1],
                            "dog_name": row[2],
                            "data": {
                                "id": row[0],
                                "t_session_number": row[1],
                                "t_dog_name": row[2],
                                "t_date": str(row[3]) if row[3] else "",
                                "t_handler": row[4] or "",
                                "t_field_support": row[5] or "",
                                "t_location": row[6] or "",
                                "t_start_time": row[7] or "",
                                "t_finish_time": row[8] or "",
                                "t_trail_age": row[9] or "",
                                "t_trail_length": row[10] or "",
                                "t_difficulty": row[11] or "",
                                "t_trail_layer": row[12] or "",
                                "t_cross_track_layer": row[13] or "",
                                "t_cross_track_age": row[14] or "",
                                "t_weather_laying": row[15] or "",
                                "t_temperature_laying": row[16] or "",
                                "t_wind_speed_laying": row[17] or "",
                                "t_wind_direction_laying": row[18] or "",
                                "t_humidity_laying": row[19] or "",
                                "t_weather_running": row[20] or "",
                                "t_temperature_running": row[21] or "",
                                "t_wind_speed_running": row[22] or "",
                                "t_wind_direction_running": row[23] or "",
                                "t_humidity_running": row[24] or "",
                                "t_start_behavior": row[25] or "",
                                "t_consistency": row[26] or "",
                                "t_head_position": row[27] or "",
                                "t_pace": row[28] or "",
                                "t_indication": row[29] or "",
                                "t_time_to_complete": row[30] or "",
                                "t_success_rate": row[31] or "",
                                "t_impression": row[32] or "",
                                "t_map_files": row[33] or "",
                                "update_time": _format_update_time(update_time),
                                "uuid": session_uuid
                            }
                        }
            except Exception as te:
                # t_training_sessions table might not exist
                # print(f"Note: Could not query t_training_sessions: {te}")
                pass
        
        # Restore original DB type
        if old_db_type != db_type:
            config.DB_TYPE = old_db_type
            database.engine.dispose()
            reload(database)
            
    except Exception as e:
        # print(f"Error getting DB sessions for sync: {e}")
        import traceback
        traceback.print_exc()
    
    return result


def sync_db_to_json(db_sessions, json_dict, json_folder, db_type):
    """
    Sync database to JSON - create/update JSON files for DB entries.
    Uses file modification time for comparison (handles manual edits).
    
    Args:
        db_sessions: Dict from get_db_sessions_for_sync()
        json_dict: Dict from scan_json_folder()
        json_folder: Path to JSON folder
        db_type: Database type
        
    Returns:
        int: Number of files created/updated
    """
    import re
    count = 0
    
    for session_uuid, db_info in db_sessions.items():
        json_info = json_dict.get(session_uuid)
        
        should_write = False
        
        if not json_info:
            # Session not in JSON - create it
            should_write = True
            # print(f"Sync: DB session {session_uuid} not in JSON, creating...")
            pass
        else:
            # Compare DB update time with JSON file modification time
            db_time = _parse_update_time(db_info["update_time"])
            # Use file modification time (handles manual edits)
            json_time = json_info.get("file_mtime") or _parse_update_time(json_info.get("update_time"))
            
            if db_time and json_time and db_time > json_time:
                should_write = True
                # print(f"Sync: DB newer than JSON file for {session_uuid}")
                pass
        
        # if should_write:
        #     try:
        #         # Load related data (terrains, responses) from DB
        #         session_data = load_full_session_from_db(
        #             db_info["session_number"], 
        #             db_info["dog_name"], 
        #             db_type
        #         )
                
        #         if session_data:
        #             # Create filename
        #             dog_name = session_data.get("dog_name", "unknown")
        #             safe_dog_name = re.sub(r'[^\w\-]', '_', dog_name)
        #             session_num = session_data.get("session_number")
        #             date_str = session_data.get("date", "").replace("-", "")
        #             filename = f"{safe_dog_name}_session_{session_num}_{date_str}.json"
                    
        #             # Add backup timestamp
        #             session_data["backup_timestamp"] = datetime.now().isoformat()
                    
        #             # Write JSON
        #             filepath = Path(json_folder) / filename
        #             with open(filepath, 'w') as f:
        #                 json.dump(session_data, f, indent=2, default=str)
                    
        #             count += 1
        #             # print(f"Sync: Wrote {filepath}")
        #             pass
                    
        #     except Exception as e:
        #         # print(f"Sync error writing JSON for {session_uuid}: {e}")
        #         pass
    
    return count


def load_full_session_from_db(session_number, dog_name, db_type):
    """Load complete session data including terrains and responses from DB."""
    try:
        import config
        from sqlalchemy import text
        
        old_db_type = config.DB_TYPE
        config.DB_TYPE = db_type
        
        from database import engine, get_connection
        from importlib import reload
        import database
        
        if old_db_type != db_type:
            engine.dispose()
            reload(database)
        
        session_data = None
        
        with database.get_connection() as conn:
            # Get main session
            result = conn.execute(
                text("""
                    SELECT id, date, handler, session_purpose, field_support, dog_name, 
                           location, search_area_size, num_subjects, handler_knowledge, 
                           weather, temperature, wind_direction, wind_speed, search_type, 
                           drive_level, subjects_found, comments, image_files, 
                           entry_type, update_time, uuid, status
                    FROM training_sessions 
                    WHERE session_number = :session_number AND dog_name = :dog_name
                """),
                {"session_number": session_number, "dog_name": dog_name}
            )
            row = result.fetchone()
            
            if row:
                session_id = row[0]
                session_data = {
                    "session_number": session_number,
                    "date": str(row[1]) if row[1] else "",
                    "handler": row[2] or "",
                    "session_purpose": row[3] or "",
                    "field_support": row[4] or "",
                    "dog_name": row[5] or "",
                    "location": row[6] or "",
                    "search_area_size": row[7] or "",
                    "num_subjects": row[8] or "",
                    "handler_knowledge": row[9] or "",
                    "weather": row[10] or "",
                    "temperature": row[11] or "",
                    "wind_direction": row[12] or "",
                    "wind_speed": row[13] or "",
                    "search_type": row[14] or "",
                    "drive_level": row[15] or "",
                    "subjects_found": row[16] or "",
                    "comments": row[17] or "",
                    "image_files": json.loads(row[18]) if row[18] else [],
                    "entry_type": row[19] or "",
                    "update_time": _format_update_time(row[20]),
                    "uuid": row[21] or "",
                    "status": row[22] or "active"
                }
                
                # Get terrains
                terrain_result = conn.execute(
                    text("SELECT terrain_name FROM selected_terrains WHERE session_id = :session_id"),
                    {"session_id": session_id}
                )
                session_data["selected_terrains"] = [r[0] for r in terrain_result]
                
                # Get subject responses
                response_result = conn.execute(
                    text("""
                        SELECT subject_number, tfr, refind 
                        FROM subject_responses 
                        WHERE session_id = :session_id 
                        ORDER BY subject_number
                    """),
                    {"session_id": session_id}
                )
                session_data["subject_responses"] = [
                    {"subject_number": r[0], "tfr": r[1] or "", "refind": r[2] or ""}
                    for r in response_result
                ]
        
        # Restore original DB type
        if old_db_type != db_type:
            config.DB_TYPE = old_db_type
            database.engine.dispose()
            reload(database)
        
        return session_data
        
    except Exception as e:
        # print(f"Error loading full session from DB: {e}")
        return None


def sync_json_to_db(json_dict, db_sessions, db_type):
    """
    Sync JSON to database - create missing or update DB entries where JSON is newer.
    Uses file modification time for comparison (handles manual edits).
    
    Returns:
        int: Number of DB records created/updated
    """
    count = 0
    
    for key, json_info in json_dict.items():
        if not json_info.get("has_uuid"):
            continue  # Skip files without UUID
        
        session_uuid = key
        db_info = db_sessions.get(session_uuid)
        
        if not db_info:
            # Session exists in JSON but NOT in DB - create it
            # print(f"Sync: Session {session_uuid} in JSON but not in DB, creating...")
            try:
                if insert_session_from_json(json_info["data"], db_type):
                    count += 1
            except Exception as e:
                # print(f"Sync error inserting session from JSON: {e}")
                pass
        else:
            # Both exist - check if JSON file is newer (use file mtime for manual edits)
            db_time = _parse_update_time(db_info["update_time"])
            # Use file modification time, falling back to update_time from JSON
            json_time = json_info.get("file_mtime") or _parse_update_time(json_info.get("update_time"))
            
            if db_time and json_time and json_time > db_time:
                # print(f"Sync: JSON file newer than DB for {session_uuid}, updating DB...")
                try:
                    if update_db_from_json(json_info["data"], db_type):
                        count += 1
                except Exception as e:
                    # print(f"Sync error updating DB from JSON: {e}")
                    pass
    
    return count
    
    return count


def insert_session_from_json(json_data, db_type):
    """Insert a new session into database from JSON data.
    
    Automatically detects whether it's an airscenting or trailing session
    based on the presence of 't_session_number' key.
    """
    # Detect session type
    is_trailing = 't_session_number' in json_data
    
    if is_trailing:
        return insert_trailing_session_from_json(json_data, db_type)
    else:
        return insert_airscenting_session_from_json(json_data, db_type)


def insert_airscenting_session_from_json(json_data, db_type):
    """Insert a new airscenting session into database from JSON data."""
    try:
        import config
        from sqlalchemy import text
        from ui_utils import get_username
        
        old_db_type = config.DB_TYPE
        config.DB_TYPE = db_type
        
        from database import engine, get_connection
        from importlib import reload
        import database
        
        if old_db_type != db_type:
            engine.dispose()
            reload(database)
        
        with database.get_connection() as conn:
            image_files = json_data.get("image_files", [])
            if isinstance(image_files, list):
                image_files_json = json.dumps(image_files)
            else:
                image_files_json = image_files or ""
            
            # Check if session already exists (by session_number + dog_name)
            check_result = conn.execute(
                text("""
                    SELECT id FROM training_sessions 
                    WHERE session_number = :session_number AND dog_name = :dog_name
                """),
                {
                    "session_number": json_data.get("session_number"),
                    "dog_name": json_data.get("dog_name")
                }
            )
            existing = check_result.fetchone()
            
            if existing:
                # Session exists with same number/dog but different UUID - update it
                # print(f"Sync: Session {json_data.get('session_number')}/{json_data.get('dog_name')} exists, updating with UUID...")
                conn.execute(
                    text("""
                        UPDATE training_sessions SET
                            date = :date,
                            handler = :handler,
                            session_purpose = :session_purpose,
                            field_support = :field_support,
                            location = :location,
                            search_area_size = :search_area_size,
                            num_subjects = :num_subjects,
                            handler_knowledge = :handler_knowledge,
                            weather = :weather,
                            temperature = :temperature,
                            wind_direction = :wind_direction,
                            wind_speed = :wind_speed,
                            search_type = :search_type,
                            drive_level = :drive_level,
                            subjects_found = :subjects_found,
                            comments = :comments,
                            image_files = :image_files,
                            entry_type = :entry_type,
                            update_time = :update_time,
                            uuid = :uuid,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE session_number = :session_number AND dog_name = :dog_name
                    """),
                    {
                        "date": json_data.get("date"),
                        "handler": json_data.get("handler"),
                        "session_purpose": json_data.get("session_purpose"),
                        "field_support": json_data.get("field_support"),
                        "location": json_data.get("location"),
                        "search_area_size": json_data.get("search_area_size"),
                        "num_subjects": json_data.get("num_subjects"),
                        "handler_knowledge": json_data.get("handler_knowledge"),
                        "weather": json_data.get("weather"),
                        "temperature": json_data.get("temperature"),
                        "wind_direction": json_data.get("wind_direction"),
                        "wind_speed": json_data.get("wind_speed"),
                        "search_type": json_data.get("search_type"),
                        "drive_level": json_data.get("drive_level"),
                        "subjects_found": json_data.get("subjects_found"),
                        "comments": json_data.get("comments"),
                        "image_files": image_files_json,
                        "entry_type": json_data.get("entry_type"),
                        "update_time": json_data.get("update_time"),
                        "uuid": json_data.get("uuid"),
                        "session_number": json_data.get("session_number"),
                        "dog_name": json_data.get("dog_name")
                    }
                )
            else:
                # Insert new session
                conn.execute(
                    text("""
                        INSERT INTO training_sessions 
                        (date, session_number, handler, session_purpose, field_support, dog_name, 
                         location, search_area_size, num_subjects, handler_knowledge, 
                         weather, temperature, wind_direction, wind_speed, search_type, 
                         drive_level, subjects_found, comments, image_files, 
                         entry_type, update_time, uuid, user_name, status)
                        VALUES (:date, :session_number, :handler, :session_purpose, :field_support, :dog_name,
                                :location, :search_area_size, :num_subjects, :handler_knowledge,
                                :weather, :temperature, :wind_direction, :wind_speed, :search_type,
                                :drive_level, :subjects_found, :comments, :image_files,
                                :entry_type, :update_time, :uuid, :user_name, :status)
                    """),
                    {
                        "date": json_data.get("date"),
                        "session_number": json_data.get("session_number"),
                        "handler": json_data.get("handler"),
                        "session_purpose": json_data.get("session_purpose"),
                        "field_support": json_data.get("field_support"),
                        "dog_name": json_data.get("dog_name"),
                        "location": json_data.get("location"),
                        "search_area_size": json_data.get("search_area_size"),
                        "num_subjects": json_data.get("num_subjects"),
                        "handler_knowledge": json_data.get("handler_knowledge"),
                        "weather": json_data.get("weather"),
                        "temperature": json_data.get("temperature"),
                        "wind_direction": json_data.get("wind_direction"),
                        "wind_speed": json_data.get("wind_speed"),
                        "search_type": json_data.get("search_type"),
                        "drive_level": json_data.get("drive_level"),
                        "subjects_found": json_data.get("subjects_found"),
                        "comments": json_data.get("comments"),
                        "image_files": image_files_json,
                        "entry_type": json_data.get("entry_type"),
                        "update_time": json_data.get("update_time"),
                        "uuid": json_data.get("uuid"),
                        "user_name": json_data.get("user_name", get_username()),
                        "status": json_data.get("status", "active")
                    }
                )
            
            conn.commit()
            # print(f"Sync: Inserted/updated airscenting session {json_data.get('session_number')} for {json_data.get('dog_name')}")
            pass
        
        # Restore original DB type
        if old_db_type != db_type:
            config.DB_TYPE = old_db_type
            database.engine.dispose()
            reload(database)
        
        return True
        
    except Exception as e:
        # print(f"Error inserting airscenting session from JSON: {e}")
        import traceback
        traceback.print_exc()
        return False


def insert_trailing_session_from_json(json_data, db_type):
    """Insert a new trailing session into database from JSON data."""
    try:
        import config
        from sqlalchemy import text
        from ui_utils import get_username
        
        old_db_type = config.DB_TYPE
        config.DB_TYPE = db_type
        
        from database import engine, get_connection
        from importlib import reload
        import database
        
        if old_db_type != db_type:
            engine.dispose()
            reload(database)
        
        with database.get_connection() as conn:
            map_files = json_data.get("t_map_files", [])
            if isinstance(map_files, list):
                map_files_json = json.dumps(map_files)
            else:
                map_files_json = map_files or ""
            
            # Check if session already exists (by t_session_number + t_dog_name)
            check_result = conn.execute(
                text("""
                    SELECT id FROM t_training_sessions 
                    WHERE t_session_number = :session_number AND t_dog_name = :dog_name
                """),
                {
                    "session_number": json_data.get("t_session_number"),
                    "dog_name": json_data.get("t_dog_name")
                }
            )
            existing = check_result.fetchone()
            
            if existing:
                # Session exists - update it
                # print(f"Sync: Trailing session {json_data.get('t_session_number')}/{json_data.get('t_dog_name')} exists, updating...")
                conn.execute(
                    text("""
                        UPDATE t_training_sessions SET
                            t_date = :t_date,
                            t_handler = :t_handler,
                            t_field_support = :t_field_support,
                            t_location = :t_location,
                            t_start_time = :t_start_time,
                            t_finish_time = :t_finish_time,
                            t_trail_age = :t_trail_age,
                            t_trail_length = :t_trail_length,
                            t_difficulty = :t_difficulty,
                            t_trail_layer = :t_trail_layer,
                            t_cross_track_layer = :t_cross_track_layer,
                            t_cross_track_age = :t_cross_track_age,
                            t_weather_laying = :t_weather_laying,
                            t_temperature_laying = :t_temperature_laying,
                            t_wind_speed_laying = :t_wind_speed_laying,
                            t_wind_direction_laying = :t_wind_direction_laying,
                            t_humidity_laying = :t_humidity_laying,
                            t_weather_running = :t_weather_running,
                            t_temperature_running = :t_temperature_running,
                            t_wind_speed_running = :t_wind_speed_running,
                            t_wind_direction_running = :t_wind_direction_running,
                            t_humidity_running = :t_humidity_running,
                            t_start_behavior = :t_start_behavior,
                            t_consistency = :t_consistency,
                            t_head_position = :t_head_position,
                            t_pace = :t_pace,
                            t_indication = :t_indication,
                            t_time_to_complete = :t_time_to_complete,
                            t_success_rate = :t_success_rate,
                            t_impression = :t_impression,
                            t_map_files = :t_map_files,
                            update_time = :update_time,
                            uuid = :uuid
                        WHERE t_session_number = :t_session_number AND t_dog_name = :t_dog_name
                    """),
                    {
                        "t_date": json_data.get("t_date"),
                        "t_handler": json_data.get("t_handler"),
                        "t_field_support": json_data.get("t_field_support"),
                        "t_location": json_data.get("t_location"),
                        "t_start_time": json_data.get("t_start_time"),
                        "t_finish_time": json_data.get("t_finish_time"),
                        "t_trail_age": json_data.get("t_trail_age"),
                        "t_trail_length": json_data.get("t_trail_length"),
                        "t_difficulty": json_data.get("t_difficulty"),
                        "t_trail_layer": json_data.get("t_trail_layer"),
                        "t_cross_track_layer": json_data.get("t_cross_track_layer"),
                        "t_cross_track_age": json_data.get("t_cross_track_age"),
                        "t_weather_laying": json_data.get("t_weather_laying"),
                        "t_temperature_laying": json_data.get("t_temperature_laying"),
                        "t_wind_speed_laying": json_data.get("t_wind_speed_laying"),
                        "t_wind_direction_laying": json_data.get("t_wind_direction_laying"),
                        "t_humidity_laying": json_data.get("t_humidity_laying"),
                        "t_weather_running": json_data.get("t_weather_running"),
                        "t_temperature_running": json_data.get("t_temperature_running"),
                        "t_wind_speed_running": json_data.get("t_wind_speed_running"),
                        "t_wind_direction_running": json_data.get("t_wind_direction_running"),
                        "t_humidity_running": json_data.get("t_humidity_running"),
                        "t_start_behavior": json_data.get("t_start_behavior"),
                        "t_consistency": json_data.get("t_consistency"),
                        "t_head_position": json_data.get("t_head_position"),
                        "t_pace": json_data.get("t_pace"),
                        "t_indication": json_data.get("t_indication"),
                        "t_time_to_complete": json_data.get("t_time_to_complete"),
                        "t_success_rate": json_data.get("t_success_rate"),
                        "t_impression": json_data.get("t_impression"),
                        "t_map_files": map_files_json,
                        "update_time": json_data.get("update_time"),
                        "uuid": json_data.get("uuid"),
                        "t_session_number": json_data.get("t_session_number"),
                        "t_dog_name": json_data.get("t_dog_name")
                    }
                )
            else:
                # Insert new session
                conn.execute(
                    text("""
                        INSERT INTO t_training_sessions 
                        (t_session_number, t_dog_name, t_date, t_handler, t_field_support,
                         t_location, t_start_time, t_finish_time, t_trail_age, t_trail_length,
                         t_difficulty, t_trail_layer, t_cross_track_layer, t_cross_track_age,
                         t_weather_laying, t_temperature_laying, t_wind_speed_laying, t_wind_direction_laying, t_humidity_laying,
                         t_weather_running, t_temperature_running, t_wind_speed_running, t_wind_direction_running, t_humidity_running,
                         t_start_behavior, t_consistency, t_head_position, t_pace, t_indication,
                         t_time_to_complete, t_success_rate, t_impression, t_map_files,
                         update_time, uuid, user_name, status)
                        VALUES (:t_session_number, :t_dog_name, :t_date, :t_handler, :t_field_support,
                                :t_location, :t_start_time, :t_finish_time, :t_trail_age, :t_trail_length,
                                :t_difficulty, :t_trail_layer, :t_cross_track_layer, :t_cross_track_age,
                                :t_weather_laying, :t_temperature_laying, :t_wind_speed_laying, :t_wind_direction_laying, :t_humidity_laying,
                                :t_weather_running, :t_temperature_running, :t_wind_speed_running, :t_wind_direction_running, :t_humidity_running,
                                :t_start_behavior, :t_consistency, :t_head_position, :t_pace, :t_indication,
                                :t_time_to_complete, :t_success_rate, :t_impression, :t_map_files,
                                :update_time, :uuid, :user_name, :status)
                    """),
                    {
                        "t_session_number": json_data.get("t_session_number"),
                        "t_dog_name": json_data.get("t_dog_name"),
                        "t_date": json_data.get("t_date"),
                        "t_handler": json_data.get("t_handler"),
                        "t_field_support": json_data.get("t_field_support"),
                        "t_location": json_data.get("t_location"),
                        "t_start_time": json_data.get("t_start_time"),
                        "t_finish_time": json_data.get("t_finish_time"),
                        "t_trail_age": json_data.get("t_trail_age"),
                        "t_trail_length": json_data.get("t_trail_length"),
                        "t_difficulty": json_data.get("t_difficulty"),
                        "t_trail_layer": json_data.get("t_trail_layer"),
                        "t_cross_track_layer": json_data.get("t_cross_track_layer"),
                        "t_cross_track_age": json_data.get("t_cross_track_age"),
                        "t_weather_laying": json_data.get("t_weather_laying"),
                        "t_temperature_laying": json_data.get("t_temperature_laying"),
                        "t_wind_speed_laying": json_data.get("t_wind_speed_laying"),
                        "t_wind_direction_laying": json_data.get("t_wind_direction_laying"),
                        "t_humidity_laying": json_data.get("t_humidity_laying"),
                        "t_weather_running": json_data.get("t_weather_running"),
                        "t_temperature_running": json_data.get("t_temperature_running"),
                        "t_wind_speed_running": json_data.get("t_wind_speed_running"),
                        "t_wind_direction_running": json_data.get("t_wind_direction_running"),
                        "t_humidity_running": json_data.get("t_humidity_running"),
                        "t_start_behavior": json_data.get("t_start_behavior"),
                        "t_consistency": json_data.get("t_consistency"),
                        "t_head_position": json_data.get("t_head_position"),
                        "t_pace": json_data.get("t_pace"),
                        "t_indication": json_data.get("t_indication"),
                        "t_time_to_complete": json_data.get("t_time_to_complete"),
                        "t_success_rate": json_data.get("t_success_rate"),
                        "t_impression": json_data.get("t_impression"),
                        "t_map_files": map_files_json,
                        "update_time": json_data.get("update_time"),
                        "uuid": json_data.get("uuid"),
                        "user_name": json_data.get("user_name", get_username()),
                        "status": json_data.get("status", "active")
                    }
                )
            
            conn.commit()
            # print(f"Sync: Inserted/updated trailing session {json_data.get('t_session_number')} for {json_data.get('t_dog_name')}")
            pass
        
        # Restore original DB type
        if old_db_type != db_type:
            config.DB_TYPE = old_db_type
            database.engine.dispose()
            reload(database)
        
        return True
        
    except Exception as e:
        # print(f"Error inserting trailing session from JSON: {e}")
        import traceback
        traceback.print_exc()
        return False


def update_db_from_json(json_data, db_type):
    """Update database record from JSON data.
    
    Automatically detects whether it's an airscenting or trailing session
    based on the presence of 't_session_number' key.
    """
    # Detect session type
    is_trailing = 't_session_number' in json_data
    
    if is_trailing:
        return update_trailing_db_from_json(json_data, db_type)
    else:
        return update_airscenting_db_from_json(json_data, db_type)


def update_airscenting_db_from_json(json_data, db_type):
    """Update airscenting database record from JSON data."""
    try:
        import config
        from sqlalchemy import text
        from ui_utils import get_username
        
        old_db_type = config.DB_TYPE
        config.DB_TYPE = db_type
        
        from database import engine, get_connection
        from importlib import reload
        import database
        
        if old_db_type != db_type:
            engine.dispose()
            reload(database)
        
        with database.get_connection() as conn:
            # Update main session
            image_files = json_data.get("image_files", [])
            if isinstance(image_files, list):
                image_files_json = json.dumps(image_files)
            else:
                image_files_json = image_files or ""
            
            conn.execute(
                text("""
                    UPDATE training_sessions SET
                        date = :date,
                        handler = :handler,
                        session_purpose = :session_purpose,
                        field_support = :field_support,
                        location = :location,
                        search_area_size = :search_area_size,
                        num_subjects = :num_subjects,
                        handler_knowledge = :handler_knowledge,
                        weather = :weather,
                        temperature = :temperature,
                        wind_direction = :wind_direction,
                        wind_speed = :wind_speed,
                        search_type = :search_type,
                        drive_level = :drive_level,
                        subjects_found = :subjects_found,
                        comments = :comments,
                        image_files = :image_files,
                        update_time = :update_time,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE uuid = :uuid
                """),
                {
                    "date": json_data.get("date"),
                    "handler": json_data.get("handler"),
                    "session_purpose": json_data.get("session_purpose"),
                    "field_support": json_data.get("field_support"),
                    "location": json_data.get("location"),
                    "search_area_size": json_data.get("search_area_size"),
                    "num_subjects": json_data.get("num_subjects"),
                    "handler_knowledge": json_data.get("handler_knowledge"),
                    "weather": json_data.get("weather"),
                    "temperature": json_data.get("temperature"),
                    "wind_direction": json_data.get("wind_direction"),
                    "wind_speed": json_data.get("wind_speed"),
                    "search_type": json_data.get("search_type"),
                    "drive_level": json_data.get("drive_level"),
                    "subjects_found": json_data.get("subjects_found"),
                    "comments": json_data.get("comments"),
                    "image_files": image_files_json,
                    "update_time": json_data.get("update_time"),
                    "uuid": json_data.get("uuid")
                }
            )
            conn.commit()
        
        # Restore original DB type
        if old_db_type != db_type:
            config.DB_TYPE = old_db_type
            database.engine.dispose()
            reload(database)
        
        return True
        
    except Exception as e:
        # print(f"Error updating airscenting DB from JSON: {e}")
        return False


def update_trailing_db_from_json(json_data, db_type):
    """Update trailing database record from JSON data."""
    try:
        import config
        from sqlalchemy import text
        
        old_db_type = config.DB_TYPE
        config.DB_TYPE = db_type
        
        from database import engine, get_connection
        from importlib import reload
        import database
        
        if old_db_type != db_type:
            engine.dispose()
            reload(database)
        
        with database.get_connection() as conn:
            # Handle map files
            map_files = json_data.get("t_map_files", [])
            if isinstance(map_files, list):
                map_files_json = json.dumps(map_files)
            else:
                map_files_json = map_files or ""
            
            conn.execute(
                text("""
                    UPDATE t_training_sessions SET
                        t_date = :t_date,
                        t_handler = :t_handler,
                        t_field_support = :t_field_support,
                        t_location = :t_location,
                        t_start_time = :t_start_time,
                        t_finish_time = :t_finish_time,
                        t_trail_age = :t_trail_age,
                        t_trail_length = :t_trail_length,
                        t_difficulty = :t_difficulty,
                        t_trail_layer = :t_trail_layer,
                        t_cross_track_layer = :t_cross_track_layer,
                        t_cross_track_age = :t_cross_track_age,
                        t_weather_laying = :t_weather_laying,
                        t_temperature_laying = :t_temperature_laying,
                        t_wind_speed_laying = :t_wind_speed_laying,
                        t_wind_direction_laying = :t_wind_direction_laying,
                        t_humidity_laying = :t_humidity_laying,
                        t_weather_running = :t_weather_running,
                        t_temperature_running = :t_temperature_running,
                        t_wind_speed_running = :t_wind_speed_running,
                        t_wind_direction_running = :t_wind_direction_running,
                        t_humidity_running = :t_humidity_running,
                        t_start_behavior = :t_start_behavior,
                        t_consistency = :t_consistency,
                        t_head_position = :t_head_position,
                        t_pace = :t_pace,
                        t_indication = :t_indication,
                        t_time_to_complete = :t_time_to_complete,
                        t_success_rate = :t_success_rate,
                        t_impression = :t_impression,
                        t_map_files = :t_map_files,
                        update_time = :update_time
                    WHERE uuid = :uuid
                """),
                {
                    "t_date": json_data.get("t_date"),
                    "t_handler": json_data.get("t_handler"),
                    "t_field_support": json_data.get("t_field_support"),
                    "t_location": json_data.get("t_location"),
                    "t_start_time": json_data.get("t_start_time"),
                    "t_finish_time": json_data.get("t_finish_time"),
                    "t_trail_age": json_data.get("t_trail_age"),
                    "t_trail_length": json_data.get("t_trail_length"),
                    "t_difficulty": json_data.get("t_difficulty"),
                    "t_trail_layer": json_data.get("t_trail_layer"),
                    "t_cross_track_layer": json_data.get("t_cross_track_layer"),
                    "t_cross_track_age": json_data.get("t_cross_track_age"),
                    "t_weather_laying": json_data.get("t_weather_laying"),
                    "t_temperature_laying": json_data.get("t_temperature_laying"),
                    "t_wind_speed_laying": json_data.get("t_wind_speed_laying"),
                    "t_wind_direction_laying": json_data.get("t_wind_direction_laying"),
                    "t_humidity_laying": json_data.get("t_humidity_laying"),
                    "t_weather_running": json_data.get("t_weather_running"),
                    "t_temperature_running": json_data.get("t_temperature_running"),
                    "t_wind_speed_running": json_data.get("t_wind_speed_running"),
                    "t_wind_direction_running": json_data.get("t_wind_direction_running"),
                    "t_humidity_running": json_data.get("t_humidity_running"),
                    "t_start_behavior": json_data.get("t_start_behavior"),
                    "t_consistency": json_data.get("t_consistency"),
                    "t_head_position": json_data.get("t_head_position"),
                    "t_pace": json_data.get("t_pace"),
                    "t_indication": json_data.get("t_indication"),
                    "t_time_to_complete": json_data.get("t_time_to_complete"),
                    "t_success_rate": json_data.get("t_success_rate"),
                    "t_impression": json_data.get("t_impression"),
                    "t_map_files": map_files_json,
                    "update_time": json_data.get("update_time"),
                    "uuid": json_data.get("uuid")
                }
            )
            conn.commit()
            # print(f"Sync: Updated trailing session {json_data.get('t_session_number')} from JSON")
            pass
        
        # Restore original DB type
        if old_db_type != db_type:
            config.DB_TYPE = old_db_type
            database.engine.dispose()
            reload(database)
        
        return True
        
    except Exception as e:
        # print(f"Error updating trailing DB from JSON: {e}")
        import traceback
        traceback.print_exc()
        return False


def validate_json_file(filepath):
    """Validate that a JSON file can be loaded without errors.
    
    Args:
        filepath: Path to the JSON file
        
    Returns:
        bool: True if valid, False otherwise
    """
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            json.load(f)
        return True
    except Exception as e:
        # print(f"JSON validation failed for {filepath}: {e}")
        return False


def sync_primary_to_secondary(primary_dict, secondary_dict, secondary_folder):
    """
    Sync primary JSON folder to secondary - copy newer/missing files.
    Only copies valid JSON files.
    Uses file modification time for comparison (handles manual edits).
    
    Returns:
        int: Number of files copied
    """
    count = 0
    
    for key, primary_info in primary_dict.items():
        secondary_info = secondary_dict.get(key)
        
        should_copy = False
        
        if not secondary_info:
            # File not in secondary
            should_copy = True
        else:
            # Use file modification time for comparison (more reliable for manual edits)
            primary_mtime = primary_info.get("file_mtime") or primary_info.get("update_time")
            secondary_mtime = secondary_info.get("file_mtime") or secondary_info.get("update_time")
            
            if primary_mtime and secondary_mtime:
                if primary_mtime > secondary_mtime:
                    should_copy = True
        
        if should_copy:
            try:
                src = primary_info["filepath"]
                
                # Validate JSON before copying
                if not validate_json_file(src):
                    # print(f"Sync: Skipping invalid JSON file: {src.name}")
                    continue
                
                dst = Path(secondary_folder) / src.name
                shutil.copy2(str(src), str(dst))
                count += 1
                # print(f"Sync: Copied to secondary: {dst.name}")
                pass
            except Exception as e:
                # print(f"Sync error copying to secondary: {e}")
                pass
    
    return count


def sync_secondary_to_primary(secondary_dict, primary_dict, primary_folder, db_type):
    """
    Sync secondary JSON folder to primary - copy newer files and update DB.
    Only copies valid JSON files.
    Uses file modification time for comparison (handles manual edits).
    
    Returns:
        int: Number of files copied
    """
    count = 0
    
    for key, secondary_info in secondary_dict.items():
        primary_info = primary_dict.get(key)
        
        should_copy = False
        
        if not primary_info:
            # File not in primary
            should_copy = True
        else:
            # Use file modification time for comparison (more reliable for manual edits)
            secondary_mtime = secondary_info.get("file_mtime") or secondary_info.get("update_time")
            primary_mtime = primary_info.get("file_mtime") or primary_info.get("update_time")
            
            if secondary_mtime and primary_mtime:
                if secondary_mtime > primary_mtime:
                    should_copy = True
        
        if should_copy:
            try:
                src = secondary_info["filepath"]
                
                # Validate JSON before copying
                if not validate_json_file(src):
                    # print(f"Sync: Skipping invalid JSON file from secondary: {src.name}")
                    continue
                
                dst = Path(primary_folder) / src.name
                shutil.copy2(str(src), str(dst))
                count += 1
                # print(f"Sync: Copied from secondary to primary: {dst.name}")
                pass
                
                # Also update DB if this file has a UUID
                if secondary_info.get("has_uuid"):
                    update_db_from_json(secondary_info["data"], db_type)
                    
            except Exception as e:
                # print(f"Sync error copying from secondary: {e}")
                pass
    
    return count


# Global sync manager instance
_sync_manager = None

def get_sync_manager():
    """Get the global sync manager instance."""
    global _sync_manager
    if _sync_manager is None:
        _sync_manager = BackupSyncManager()
    return _sync_manager


# ============================================================================
# EXCEL EXPORT FUNCTIONALITY
# ============================================================================

# Area Search UI field names mapping (DB column -> UI name)
AREA_SEARCH_FIELD_MAPPING = {
    'session_number': 'Session #',
    'date': 'Date',
    'handler': 'Handler',
    'session_purpose': 'Session Purpose',
    'field_support': 'Field Support',
    'dog_name': 'Dog',
    'location': 'Location',
    'search_area_size': 'Search Area (Acres)',
    'num_subjects': 'Number of Subjects',
    'handler_knowledge': 'Handler Knowledge',
    'weather': 'Weather',
    'temperature': 'Temperature',
    'wind_direction': 'Wind Direction',
    'wind_speed': 'Wind Speed',
    'search_type': 'Search Type',
    'drive_level': 'Drive Level',
    'subjects_found': 'Subjects Found',
    'a_percent_searched': 'Percent Searched',
    'start_time': 'Start Time',
    'finish_time': 'Finish Time',
    'comments': 'Comments',
    'selected_terrains': 'Terrain Types',
    'selected_purposes': 'Session Purposes',
    'subject_responses': 'Subject Responses',
    'image_files': 'Image Files',
    'uuid': 'UUID',
    'status': 'Status',
}

# Trailing UI field names mapping (DB column -> UI name)
TRAILING_FIELD_MAPPING = {
    't_session_number': 'Session #',
    't_date': 'Date',
    't_handler': 'Handler',
    't_field_support': 'Field Support',
    't_dog_name': 'Dog',
    't_location': 'Location',
    't_start_time': 'Start Time',
    't_finish_time': 'Finish Time',
    't_trail_age': 'Trail Age',
    't_trail_length': 'Trail Length',
    't_difficulty': 'Difficulty',
    't_trail_layer': 'Trail Layer',
    't_cross_track_layer': 'Cross Track Layer',
    't_cross_track_age': 'Cross Track Age',
    't_weather_laying': 'Weather (Laying)',
    't_temperature_laying': 'Temperature (Laying)',
    't_wind_speed_laying': 'Wind Speed (Laying)',
    't_wind_direction_laying': 'Wind Direction (Laying)',
    't_humidity_laying': 'Humidity (Laying)',
    't_weather_running': 'Weather (Running)',
    't_temperature_running': 'Temperature (Running)',
    't_wind_speed_running': 'Wind Speed (Running)',
    't_wind_direction_running': 'Wind Direction (Running)',
    't_humidity_running': 'Humidity (Running)',
    't_start_behavior': 'Start Behavior',
    't_consistency': 'Consistency',
    't_head_position': 'Head Position',
    't_pace': 'Pace',
    't_indication': 'Indication',
    't_time_to_complete': 'Time to Complete',
    't_success_rate': 'Success Rate',
    't_impression': 'Impression',
    't_selected_terrains': 'Terrain Types',
    't_selected_purposes': 'Session Purposes',
    't_distractions': 'Distractions',
    't_map_files': 'Map Files',
    'uuid': 'UUID',
    'status': 'Status',
}


def export_sessions_to_excel(db_type, json_folder, session_type='airscent'):
    """
    Export all sessions to Excel file with each dog on a separate sheet.
    
    Args:
        db_type: Database type (sqlite, postgres, etc.)
        json_folder: Path to the JSON folder where Excel will be saved
        session_type: 'airscent' or 'trailing'
        
    Returns:
        tuple: (success: bool, message: str, filepath: str or None)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False, "openpyxl not installed. Run: pip install openpyxl", None
    
    try:
        import config
        from sqlalchemy import text
        
        old_db_type = config.DB_TYPE
        config.DB_TYPE = db_type
        
        from database import engine, get_connection
        from importlib import reload
        import database
        
        if old_db_type != db_type:
            engine.dispose()
            reload(database)
        
        # Create workbook
        wb = Workbook()
        # Remove default sheet - we'll create dog-specific sheets
        if wb.active:
            wb.remove(wb.active)
        
        # Styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        sessions_exported = 0
        
        with database.get_connection() as conn:
            if session_type == 'trailing':
                # Get all trailing sessions grouped by dog
                field_mapping = TRAILING_FIELD_MAPPING
                table_name = 't_training_sessions'
                dog_column = 't_dog_name'
                session_num_column = 't_session_number'
                
                # Get unique dogs
                dogs_result = conn.execute(text(f"""
                    SELECT DISTINCT {dog_column} FROM {table_name}
                    WHERE {dog_column} IS NOT NULL AND {dog_column} != ''
                    ORDER BY {dog_column}
                """))
                dogs = [row[0] for row in dogs_result]
                
                for dog_name in dogs:
                    # Create sheet for this dog
                    safe_sheet_name = dog_name[:31] if len(dog_name) > 31 else dog_name
                    safe_sheet_name = safe_sheet_name.replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace('[', '-').replace(']', '-')
                    ws = wb.create_sheet(title=safe_sheet_name)
                    
                    # Get sessions for this dog
                    sessions_result = conn.execute(text(f"""
                        SELECT * FROM {table_name}
                        WHERE {dog_column} = :dog_name
                        ORDER BY {session_num_column}
                    """), {"dog_name": dog_name})
                    
                    sessions = sessions_result.fetchall()
                    columns = list(sessions_result.keys())
                    
                    # Add columns for related data
                    extra_columns = ['terrains', 'purposes', 'distractions']
                    all_columns = columns + extra_columns
                    
                    # Write headers using UI names
                    headers = []
                    for col in all_columns:
                        if col == 'terrains':
                            headers.append('Terrain Types')
                        elif col == 'purposes':
                            headers.append('Session Purposes')
                        elif col == 'distractions':
                            headers.append('Distractions')
                        else:
                            ui_name = field_mapping.get(col, col.replace('_', ' ').title())
                            headers.append(ui_name)
                    
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # Write data
                    for row_idx, session in enumerate(sessions, 2):
                        session_id = session[0]  # id is first column
                        
                        # Write main session data
                        for col_idx, value in enumerate(session, 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if value is not None:
                                if isinstance(value, (list, dict)):
                                    cell.value = json.dumps(value)
                                else:
                                    cell.value = str(value)
                            cell.border = thin_border
                        
                        # Fetch and write related data
                        base_col = len(columns) + 1
                        
                        # Terrains (comma separated only)
                        try:
                            terrains_result = conn.execute(text(
                                "SELECT terrain_name FROM t_selected_terrains WHERE t_session_id = :sid"
                            ), {"sid": session_id})
                            terrains = [r[0] for r in terrains_result]
                            cell = ws.cell(row=row_idx, column=base_col)
                            cell.value = ','.join(terrains) if terrains else ''
                            cell.border = thin_border
                        except:
                            pass
                        
                        # Purposes (comma separated only)
                        try:
                            purposes_result = conn.execute(text(
                                "SELECT purpose_name FROM t_selected_purposes WHERE t_session_id = :sid"
                            ), {"sid": session_id})
                            purposes = [r[0] for r in purposes_result]
                            cell = ws.cell(row=row_idx, column=base_col + 1)
                            cell.value = ','.join(purposes) if purposes else ''
                            cell.border = thin_border
                        except:
                            pass
                        
                        # Distractions (distraction1:response1,distraction2:response2 format)
                        try:
                            distractions_result = conn.execute(text(
                                "SELECT distraction_data FROM t_distractions WHERE t_session_id = :sid"
                            ), {"sid": session_id})
                            distractions = []
                            for r in distractions_result:
                                try:
                                    d = json.loads(r[0])
                                    dtype = d.get('type', '')
                                    dresp = d.get('response', '')
                                    distractions.append(f"{dtype}:{dresp}")
                                except:
                                    pass
                            cell = ws.cell(row=row_idx, column=base_col + 2)
                            cell.value = ','.join(distractions) if distractions else ''
                            cell.border = thin_border
                        except:
                            pass
                        
                        sessions_exported += 1
                    
                    # Adjust column widths
                    for col_idx in range(1, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            else:  # airscent
                field_mapping = AREA_SEARCH_FIELD_MAPPING
                table_name = 'training_sessions'
                dog_column = 'dog_name'
                session_num_column = 'session_number'
                
                # Get unique dogs
                dogs_result = conn.execute(text(f"""
                    SELECT DISTINCT {dog_column} FROM {table_name}
                    WHERE {dog_column} IS NOT NULL AND {dog_column} != ''
                    ORDER BY {dog_column}
                """))
                dogs = [row[0] for row in dogs_result]
                
                for dog_name in dogs:
                    # Create sheet for this dog
                    safe_sheet_name = dog_name[:31] if len(dog_name) > 31 else dog_name
                    safe_sheet_name = safe_sheet_name.replace('/', '-').replace('\\', '-').replace('*', '-').replace('?', '-').replace('[', '-').replace(']', '-')
                    ws = wb.create_sheet(title=safe_sheet_name)
                    
                    # Get sessions for this dog
                    sessions_result = conn.execute(text(f"""
                        SELECT * FROM {table_name}
                        WHERE {dog_column} = :dog_name
                        ORDER BY {session_num_column}
                    """), {"dog_name": dog_name})
                    
                    sessions = sessions_result.fetchall()
                    columns = list(sessions_result.keys())
                    
                    # Add columns for related data
                    extra_columns = ['terrains', 'purposes', 'subject_responses']
                    all_columns = columns + extra_columns
                    
                    # Write headers using UI names
                    headers = []
                    for col in all_columns:
                        if col == 'terrains':
                            headers.append('Terrain Types')
                        elif col == 'purposes':
                            headers.append('Session Purposes')
                        elif col == 'subject_responses':
                            headers.append('Subject Responses')
                        else:
                            ui_name = field_mapping.get(col, col.replace('_', ' ').title())
                            headers.append(ui_name)
                    
                    for col_idx, header in enumerate(headers, 1):
                        cell = ws.cell(row=1, column=col_idx, value=header)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # Write data
                    for row_idx, session in enumerate(sessions, 2):
                        session_id = session[0]  # id is first column
                        
                        # Write main session data
                        for col_idx, value in enumerate(session, 1):
                            cell = ws.cell(row=row_idx, column=col_idx)
                            if value is not None:
                                if isinstance(value, (list, dict)):
                                    cell.value = json.dumps(value)
                                else:
                                    cell.value = str(value)
                            cell.border = thin_border
                        
                        # Fetch and write related data
                        base_col = len(columns) + 1
                        
                        # Terrains (comma separated)
                        try:
                            terrains_result = conn.execute(text(
                                "SELECT terrain_name FROM selected_terrains WHERE session_id = :sid"
                            ), {"sid": session_id})
                            terrains = [r[0] for r in terrains_result]
                            cell = ws.cell(row=row_idx, column=base_col)
                            cell.value = ','.join(terrains) if terrains else ''
                            cell.border = thin_border
                        except:
                            pass
                        
                        # Purposes (comma separated)
                        try:
                            purposes_result = conn.execute(text(
                                "SELECT purpose_name FROM a_selected_purposes WHERE session_id = :sid"
                            ), {"sid": session_id})
                            purposes = [r[0] for r in purposes_result]
                            cell = ws.cell(row=row_idx, column=base_col + 1)
                            cell.value = ','.join(purposes) if purposes else ''
                            cell.border = thin_border
                        except:
                            pass
                        
                        # Subject responses (subject:tfr:refind format)
                        try:
                            responses_result = conn.execute(text(
                                "SELECT subject_number, tfr, refind FROM subject_responses WHERE session_id = :sid ORDER BY subject_number"
                            ), {"sid": session_id})
                            responses = [f"{r[0]}:{r[1] or ''}:{r[2] or ''}" for r in responses_result]
                            cell = ws.cell(row=row_idx, column=base_col + 2)
                            cell.value = ','.join(responses) if responses else ''
                            cell.border = thin_border
                        except:
                            pass
                        
                        sessions_exported += 1
                    
                    # Adjust column widths
                    for col_idx in range(1, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # Restore original DB type
        if old_db_type != db_type:
            config.DB_TYPE = old_db_type
            database.engine.dispose()
            reload(database)
        
        # Ensure at least one sheet exists
        if not wb.sheetnames:
            wb.create_sheet(title="No Data")
            ws = wb.active
            ws.cell(row=1, column=1, value="No sessions found")
        
        # Save Excel file
        json_path = Path(json_folder)
        if session_type == 'trailing':
            filename = f"trailing_sessions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            filename = f"area_search_sessions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = json_path / filename
        wb.save(str(filepath))
        
        return True, f"Exported {sessions_exported} sessions to Excel", str(filepath)
        
    except Exception as e:
        # Restore original DB type on error
        try:
            import config
            import database
            from importlib import reload
            if 'old_db_type' in dir() and old_db_type != db_type:
                config.DB_TYPE = old_db_type
                database.engine.dispose()
                reload(database)
        except:
            pass
        
        import traceback
        traceback.print_exc()
        return False, f"Excel export error: {e}", None


def restore_sessions_from_excel(excel_filepath, db_type, session_type='airscent'):
    """
    Restore sessions from an Excel file.
    Each sheet represents a dog, with sessions as rows.
    
    Args:
        excel_filepath: Path to the Excel file
        db_type: Database type (sqlite, postgres, etc.)
        session_type: 'airscent' or 'trailing'
        
    Returns:
        tuple: (success: bool, message: str, sessions_restored: int)
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False, "openpyxl not installed. Run: pip install openpyxl", 0
    
    try:
        import config
        from sqlalchemy import text
        
        old_db_type = config.DB_TYPE
        config.DB_TYPE = db_type
        
        from database import engine, get_connection
        from importlib import reload
        import database
        
        if old_db_type != db_type:
            engine.dispose()
            reload(database)
        
        # Load workbook
        wb = load_workbook(str(excel_filepath))
        
        sessions_restored = 0
        errors = []
        
        # Reverse the field mapping to go from UI name -> DB column
        if session_type == 'trailing':
            field_mapping = {v: k for k, v in TRAILING_FIELD_MAPPING.items()}
            table_name = 't_training_sessions'
            session_num_col = 't_session_number'
            dog_name_col = 't_dog_name'
        else:
            field_mapping = {v: k for k, v in AREA_SEARCH_FIELD_MAPPING.items()}
            table_name = 'training_sessions'
            session_num_col = 'session_number'
            dog_name_col = 'dog_name'
        
        with database.get_connection() as conn:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        # Map UI name back to DB column
                        db_col = field_mapping.get(cell.value, cell.value.lower().replace(' ', '_'))
                        headers.append(db_col)
                    else:
                        headers.append(None)
                
                # Skip sheets with no data
                if ws.max_row < 2:
                    continue
                
                # Process each row (starting from row 2)
                for row in ws.iter_rows(min_row=2):
                    row_data = {}
                    terrains_data = None
                    purposes_data = None
                    distractions_data = None
                    responses_data = None
                    
                    for idx, cell in enumerate(row):
                        if idx < len(headers) and headers[idx]:
                            col_name = headers[idx]
                            # Check for special columns (handle both mapped and unmapped names)
                            if col_name in ('terrain_types', 'terrains', 'selected_terrains', 't_selected_terrains'):
                                terrains_data = cell.value
                            elif col_name in ('session_purposes', 'purposes', 'selected_purposes', 't_selected_purposes'):
                                purposes_data = cell.value
                            elif col_name in ('distractions', 't_distractions'):
                                distractions_data = cell.value
                            elif col_name == 'subject_responses':
                                responses_data = cell.value
                            else:
                                row_data[col_name] = cell.value
                    
                    if not row_data:
                        continue
                    
                    try:
                        # Build insert statement based on session type
                        if session_type == 'trailing':
                            # Insert trailing session
                            columns = ['t_session_number', 't_date', 't_handler', 't_field_support',
                                       't_dog_name', 't_location', 't_start_time', 't_finish_time',
                                       't_trail_age', 't_trail_length', 't_difficulty', 't_trail_layer',
                                       't_cross_track_layer', 't_cross_track_age',
                                       't_weather_laying', 't_temperature_laying', 't_wind_speed_laying',
                                       't_wind_direction_laying', 't_humidity_laying',
                                       't_weather_running', 't_temperature_running', 't_wind_speed_running',
                                       't_wind_direction_running', 't_humidity_running',
                                       't_start_behavior', 't_consistency', 't_head_position', 't_pace',
                                       't_indication', 't_time_to_complete', 't_success_rate', 't_impression',
                                       't_map_files', 'uuid', 'status']
                        else:
                            columns = ['session_number', 'date', 'handler', 'session_purpose',
                                       'field_support', 'dog_name', 'location', 'search_area_size',
                                       'num_subjects', 'handler_knowledge', 'weather', 'temperature',
                                       'wind_direction', 'wind_speed', 'search_type', 'drive_level',
                                       'subjects_found', 'a_percent_searched', 'start_time', 'finish_time',
                                       'comments', 'image_files', 'uuid', 'status']
                        
                        # Filter to only columns that exist in row_data
                        insert_columns = [c for c in columns if c in row_data and row_data[c] is not None]
                        
                        if not insert_columns:
                            continue
                        
                        # Build VALUES clause with parameters
                        placeholders = ', '.join([f':{c}' for c in insert_columns])
                        col_list = ', '.join(insert_columns)
                        
                        # Build params dict
                        params = {c: row_data[c] for c in insert_columns}
                        
                        # Insert the session
                        conn.execute(text(f"INSERT INTO {table_name} ({col_list}) VALUES ({placeholders})"), params)
                        
                        # Get the session ID for inserting related data
                        session_num = row_data.get(session_num_col)
                        dog_name = row_data.get(dog_name_col)
                        
                        if session_num and dog_name:
                            result = conn.execute(text(
                                f"SELECT id FROM {table_name} WHERE {session_num_col} = :snum AND {dog_name_col} = :dname"
                            ), {"snum": session_num, "dname": dog_name})
                            row_result = result.fetchone()
                            if row_result:
                                session_id = row_result[0]
                                
                                # Insert related data
                                if session_type == 'trailing':
                                    # Terrains (comma separated)
                                    if terrains_data:
                                        for terrain in terrains_data.split(','):
                                            terrain = terrain.strip()
                                            if terrain:
                                                conn.execute(text(
                                                    "INSERT INTO t_selected_terrains (t_session_id, terrain_name) VALUES (:sid, :name)"
                                                ), {"sid": session_id, "name": terrain})
                                    
                                    # Purposes (comma separated)
                                    if purposes_data:
                                        for purpose in purposes_data.split(','):
                                            purpose = purpose.strip()
                                            if purpose:
                                                conn.execute(text(
                                                    "INSERT INTO t_selected_purposes (t_session_id, purpose_name) VALUES (:sid, :name)"
                                                ), {"sid": session_id, "name": purpose})
                                    
                                    # Distractions (type:response,type:response format)
                                    if distractions_data:
                                        for distraction in distractions_data.split(','):
                                            distraction = distraction.strip()
                                            if distraction:
                                                parts = distraction.split(':', 1)
                                                dtype = parts[0].strip()
                                                dresp = parts[1].strip() if len(parts) > 1 else ''
                                                distraction_json = json.dumps({"type": dtype, "response": dresp})
                                                conn.execute(text(
                                                    "INSERT INTO t_distractions (t_session_id, distraction_data) VALUES (:sid, :data)"
                                                ), {"sid": session_id, "data": distraction_json})
                                
                                else:  # airscent
                                    # Terrains (comma separated)
                                    if terrains_data:
                                        for terrain in terrains_data.split(','):
                                            terrain = terrain.strip()
                                            if terrain:
                                                conn.execute(text(
                                                    "INSERT INTO selected_terrains (session_id, terrain_name) VALUES (:sid, :name)"
                                                ), {"sid": session_id, "name": terrain})
                                    
                                    # Purposes (comma separated)
                                    if purposes_data:
                                        for purpose in purposes_data.split(','):
                                            purpose = purpose.strip()
                                            if purpose:
                                                conn.execute(text(
                                                    "INSERT INTO a_selected_purposes (session_id, purpose_name) VALUES (:sid, :name)"
                                                ), {"sid": session_id, "name": purpose})
                                    
                                    # Subject responses (subject:tfr:refind format)
                                    if responses_data:
                                        for response in responses_data.split(','):
                                            response = response.strip()
                                            if response:
                                                parts = response.split(':')
                                                if len(parts) >= 1:
                                                    try:
                                                        subj_num = int(parts[0].strip())
                                                        tfr = parts[1].strip() if len(parts) > 1 else ''
                                                        refind = parts[2].strip() if len(parts) > 2 else ''
                                                        conn.execute(text(
                                                            "INSERT INTO subject_responses (session_id, subject_number, tfr, refind) VALUES (:sid, :snum, :tfr, :refind)"
                                                        ), {"sid": session_id, "snum": subj_num, "tfr": tfr, "refind": refind})
                                                    except ValueError:
                                                        pass  # Skip invalid subject numbers
                        
                        sessions_restored += 1
                        
                    except Exception as e:
                        error_msg = str(e)
                        if "UNIQUE constraint" in error_msg or "duplicate key" in error_msg:
                            # Session already exists - skip (data was already cleared before restore)
                            errors.append(f"Duplicate session skipped: {row_data.get(session_num_col)}")
                        else:
                            errors.append(f"Row insert error: {e}")
            
            conn.commit()
        
        # Restore original DB type
        if old_db_type != db_type:
            config.DB_TYPE = old_db_type
            database.engine.dispose()
            reload(database)
        
        if errors:
            return True, f"Restored {sessions_restored} sessions with {len(errors)} errors", sessions_restored
        return True, f"Successfully restored {sessions_restored} sessions", sessions_restored
        
    except Exception as e:
        # Restore original DB type on error
        try:
            import config
            import database
            from importlib import reload
            if 'old_db_type' in dir() and old_db_type != db_type:
                config.DB_TYPE = old_db_type
                database.engine.dispose()
                reload(database)
        except:
            pass
        
        import traceback
        traceback.print_exc()
        return False, f"Excel restore error: {e}", 0


def export_all_sessions_to_excel(db_type, primary_folder, secondary_folder=None):
    """
    Export all sessions (both area search and trailing) to Excel files
    in both primary and secondary folders.
    
    Args:
        db_type: Database type
        primary_folder: Primary folder path (either dedicated Excel folder or primary storage)
        secondary_folder: Optional secondary backup folder path (only used if primary is not dedicated Excel folder)
        
    Returns:
        tuple: (success: bool, message: str)
    """
    results = []
    
    # Ensure folder exists
    excel_path = Path(primary_folder)
    excel_path.mkdir(parents=True, exist_ok=True)
    
    # Export area search sessions
    success, msg, filepath = export_sessions_to_excel(db_type, str(excel_path), 'airscent')
    if success:
        results.append(f"Area Search: {msg}")
        
        # Copy to secondary if exists (for JSON folder fallback mode)
        if secondary_folder and filepath:
            secondary_path = Path(secondary_folder)
            secondary_path.mkdir(parents=True, exist_ok=True)
            try:
                shutil.copy2(filepath, str(secondary_path / Path(filepath).name))
            except Exception as e:
                results.append(f"Warning: Could not copy to secondary: {e}")
    else:
        results.append(f"Area Search export failed: {msg}")
    
    # Export trailing sessions
    success, msg, filepath = export_sessions_to_excel(db_type, str(excel_path), 'trailing')
    if success:
        results.append(f"Trailing: {msg}")
        
        # Copy to secondary if exists (for JSON folder fallback mode)
        if secondary_folder and filepath:
            secondary_path = Path(secondary_folder)
            secondary_path.mkdir(parents=True, exist_ok=True)
            try:
                shutil.copy2(filepath, str(secondary_path / Path(filepath).name))
            except Exception as e:
                results.append(f"Warning: Could not copy to secondary: {e}")
    else:
        results.append(f"Trailing export failed: {msg}")
    
    return True, "; ".join(results)
